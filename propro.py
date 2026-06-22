import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import os
import re 
import tempfile 
from fpdf import FPDF 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Impor library Machine Learning untuk fitur AI
try:
    from sklearn.cluster import KMeans
    import numpy as np
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False

# --- KONFIGURASI HALAMAN ---
st.set_page_config(page_title="Sistem Analisis Profiling Siswa", layout="wide", initial_sidebar_state="expanded")

# --- UI/UX REFINEMENT: CUSTOM CSS UNTUK MENU SEGMENTED & CLEAN LOOK ---
st.markdown("""
<style>
    /* Mengubah st.radio menjadi tombol segmented horizontal yang modern */
    div[role="radiogroup"] {
        display: flex;
        flex-direction: row;
        flex-wrap: wrap;
        gap: 10px;
        justify-content: center;
    }
    div[role="radiogroup"] > label {
        background-color: #f1f5f9;
        padding: 10px 20px;
        border-radius: 30px;
        border: 1px solid #e2e8f0;
        cursor: pointer;
        transition: all 0.2s ease-in-out;
    }
    div[role="radiogroup"] > label:hover {
        background-color: #e2e8f0;
    }
    div[role="radiogroup"] > label > div:first-child {
        display: none; /* Sembunyikan lingkaran radio button */
    }
    div[role="radiogroup"] > label > div:last-child {
        font-weight: 600;
        color: #334155;
        margin-left: 0px;
    }
    /* Warna saat radio button aktif/dipilih */
    div[role="radiogroup"] > label[data-baseweb="radio"]:has(input:checked) {
        background-color: #3b82f6;
        border-color: #2563eb;
    }
    div[role="radiogroup"] > label[data-baseweb="radio"]:has(input:checked) > div:last-child {
        color: white;
    }
    /* Mempercantik kotak metrik */
    div[data-testid="metric-container"] {
        background-color: #ffffff;
        border: 1px solid #e2e8f0;
        padding: 15px;
        border-radius: 12px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    }
</style>
""", unsafe_allow_html=True)

# --- KEEP-ALIVE ---
keep_alive_js = """
<script>
(function() {
    function keepAlive() {
        fetch(window.location.href, {method: 'HEAD', cache: 'no-cache'})
            .catch(function() {});
    }
    setInterval(keepAlive, 45000);
})();
</script>
"""
st.components.v1.html(keep_alive_js, height=0)

if 'nama_kelas' not in st.session_state:
    st.session_state.nama_kelas = "Umum"

# Judul Dinamis
st.title(f"🎓 Pro-Edu Analytics: Smart Profiling {st.session_state.nama_kelas}")
st.markdown("Integrasi AI Prediktif, Personalisasi Jalur Belajar, Analisis Kedisiplinan, & Evaluasi Mata Pelajaran.")

# --- SIDEBAR: UPLOAD DATA UTAMA ---
with st.sidebar:
    st.header("📂 Setup Database Kelas")
    st.info("💡 Langkah 1: Unggah file Leger Nilai hasil unduhan e-Rapor.")
    file_leger = st.file_uploader("Upload Leger Nilai Utama:", type=['ods', 'xlsx', 'xls', 'csv'])

if not file_leger:
    st.warning("👈 Menunggu data... Silakan unggah file Leger Nilai di sidebar sebelah kiri.")
    st.stop()

# --- FUNGSI GLOBAL ---
def clean_pdf_text(text):
    return re.sub(r'[^\x00-\x7F]+', '', str(text)).strip()

def get_ekskul_bonus(predikat):
    pred = str(predikat).strip().upper()
    if pred == 'A': return 5   
    if pred == 'B': return 2   
    return 0

def generate_template_absensi(nama_siswa_list, kelas_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Absensi Ekskul"
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    info_rows = [
        ["TEMPLATE DATA ABSENSI & EKSTRAKURIKULER - BERURUTAN MUDAH DIISI"],
        ["Nama Sekolah:", "Sekolah Menengah"],
        ["Kelas:", f"{kelas_name}"],
        ["Semester:", ""],
        ["Tahun Pelajaran:", ""],
        [""],
        ["(Isi data mulai baris ke-8. Jangan merubah susunan tajuk di baris ke-8!)"],
    ]
    for r, row_data in enumerate(info_rows, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1: cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = center
    
    row8_labels = {1: "No", 2: "Nama Siswa", 3: "NIS", 4: "Sakit", 5: "Izin", 6: "Alpa", 7: "Pramuka", 8: "Kesenian", 9: "Olahraga"}
    for col_num, label in row8_labels.items():
        cell = ws.cell(row=8, column=col_num, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    
    for i, nama in enumerate(nama_siswa_list, 1):
        row_num = 8 + i
        ws.cell(row=row_num, column=1, value=i).alignment = center
        ws.cell(row=row_num, column=2, value=nama)
        ws.cell(row=row_num, column=3, value="").number_format = '@'  
        ws.cell(row=row_num, column=4, value=0).alignment = center  
        ws.cell(row=row_num, column=5, value=0).alignment = center  
        ws.cell(row=row_num, column=6, value=0).alignment = center  
        ws.cell(row=row_num, column=7, value="-").alignment = center  
        ws.cell(row=row_num, column=8, value="-").alignment = center  
        ws.cell(row=row_num, column=9, value="-").alignment = center  
        for col_set in range(1, 10):
            cell = ws.cell(row=row_num, column=col_set)
            cell.border = border
            if i % 2 == 0: cell.fill = alt_fill
    
    note_row = 8 + len(nama_siswa_list) + 2
    ws.cell(row=note_row, column=1, value="PANDUAN PENGISIAN MANUSIAWI (PENTING):").font = Font(bold=True, color="C00000")
    ws.cell(row=note_row+1, column=1, value="- Kolom Sakit, Izin, Alpa WAJIB diisi dengan angka (jumlah hari). Jika tidak ada, isi dengan 0.")
    ws.cell(row=note_row+2, column=1, value="- Kolom Ekskul diisi dengan predikat huruf kapital: A, B, C, atau strip (-) jika tidak ikut.")
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    for col_letter in ['D','E','F','G','H','I']: ws.column_dimensions[col_letter].width = 12
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generate_template_bk(nama_siswa_list, kelas_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Catatan BK"
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.cell(row=1, column=1, value=f"TEMPLATE REKAPITULASI DOSSIER & CATATAN BIMBINGAN KONSELING (BK)").font = Font(bold=True, size=12, color="7030A0")
    ws.cell(row=2, column=1, value=f"Kelas: {kelas_name}")
    ws.cell(row=3, column=1, value="(Isi data mulai baris ke-8. Kolom indeks penting agar sinkronisasi sesuai kata kunci)")
    
    row8_labels = {1: "No", 2: "Nama Siswa", 3: "NIS", 4: "Jml Pelanggaran", 11: "Catatan BK"}
    for col_num, label in row8_labels.items():
        cell = ws.cell(row=8, column=col_num, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        
    for i, nama in enumerate(nama_siswa_list, 1):
        row_num = 8 + i
        ws.cell(row=row_num, column=1, value=i).alignment = center
        ws.cell(row=row_num, column=2, value=nama)
        ws.cell(row=row_num, column=3, value="").number_format = '@'
        ws.cell(row=row_num, column=4, value=0).alignment = center
        ws.cell(row=row_num, column=11, value="-")
        for c in [1, 2, 3, 4, 11]: ws.cell(row=row_num, column=c).border = border
            
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['K'].width = 50
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generate_student_pdf(siswa_data, mapel_list, rank, total_siswa, rek_urut, mapel_urut, kelas_name, absensi_synced):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 8, "LAPORAN ANALISIS PROFILING SISWA", ln=True, align="C")
    pdf.set_font("Helvetica", "I", 10)
    pdf.cell(0, 5, f"Sistem Analisis Pro-Edu Analytics - {kelas_name}", ln=True, align="C")
    pdf.line(10, 25, 200, 25)
    pdf.ln(5)
    
    nama_clean = clean_pdf_text(siswa_data.get('Nama', ''))
    nis_clean = f"{clean_pdf_text(siswa_data.get('NISN', ''))} / {clean_pdf_text(siswa_data.get('NIS', ''))}"
    badge_clean = clean_pdf_text(siswa_data.get('Badge', ''))
    rata_rata = siswa_data.get('Rata-rata', 0)
    tot_skor = f"{siswa_data.get('Total Nilai', 0):.0f}"
    tren = siswa_data.get('Tren_Belajar', 0)
    
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, "A. PROFIL AKADEMIK & METRIK UTAMA", ln=True)
    pdf.set_font("Helvetica", "", 10)
    
    y_start_info = pdf.get_y()
    pdf.cell(30, 5, "Nama Siswa", 0, 0)
    pdf.cell(80, 5, f": {nama_clean}", 0, True)
    pdf.cell(30, 5, "NISN / NIS", 0, 0)
    pdf.cell(80, 5, f": {nis_clean}", 0, True)
    pdf.cell(30, 5, "Peringkat", 0, 0)
    pdf.cell(80, 5, f": {rank} dari {total_siswa} Siswa", 0, True)
    pdf.cell(30, 5, "Status", 0, 0)
    pdf.cell(80, 5, f": {badge_clean}", 0, True)
    
    pdf.set_xy(130, y_start_info)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(60, 5, f"Rata-rata: {rata_rata}", border=1, ln=True, align="C")
    pdf.set_xy(130, pdf.get_y() + 1)
    pdf.cell(60, 5, f"Total Skor: {tot_skor}", border=1, ln=True, align="C")
    pdf.set_xy(130, pdf.get_y() + 1)
    pdf.cell(60, 5, f"Tren Belajar: {tren} Poin", border=1, ln=True, align="C")
    pdf.ln(10)
    
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, "B. PETA KOMPETENSI & JALUR BELAJAR", ln=True)
    y_radar = pdf.get_y()
    
    categories = mapel_list
    values = [siswa_data.get(m, 0) for m in mapel_list]
    fig_radar = go.Figure()
    fig_radar.add_trace(go.Scatterpolar(r=values, theta=categories, fill='toself', line_color='#1f77b4'))
    fig_radar.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])), showlegend=False, margin=dict(l=50, r=50, t=35, b=35))
    
    tmp_radar = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    tmp_radar_name = tmp_radar.name
    tmp_radar.close() 
    try:
        fig_radar.write_image(tmp_radar_name, width=550, height=550)
        pdf.image(tmp_radar_name, x=10, y=y_radar, w=90)
    finally:
        if os.path.exists(tmp_radar_name): os.remove(tmp_radar_name)
        
    pdf.set_xy(110, y_radar + 5)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 5, "Rekomendasi Penjurusan Lanjutan:", ln=True)
    pdf.set_font("Helvetica", "", 10)
    if rek_urut and len(rek_urut) >= 2:
        pdf.set_x(110); pdf.cell(0, 5, f"1. {clean_pdf_text(rek_urut[0][0])}", ln=True)
        pdf.set_x(110); pdf.cell(0, 5, f"2. {clean_pdf_text(rek_urut[1][0])}", ln=True)
    
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_x(110); pdf.cell(0, 5, "*(Dihitung berbasis Teori Kecerdasan Hibrida)*", ln=True)
    pdf.ln(2)
    
    pdf.set_x(110); pdf.set_font("Helvetica", "B", 10); pdf.cell(0, 5, "Kekuatan Utama Belajar:", ln=True)
    pdf.set_font("Helvetica", "", 10)
    for m in mapel_urut[:3]: pdf.set_x(110); pdf.cell(0, 5, f"- {clean_pdf_text(m[0])} ({m[1]})", ln=True)
    pdf.ln(3)
    
    pdf.set_x(110); pdf.set_font("Helvetica", "B", 10); pdf.cell(0, 5, "Area Fokus Perbaikan:", ln=True)
    pdf.set_font("Helvetica", "", 10)
    for m in mapel_urut[-3:]: pdf.set_x(110); pdf.cell(0, 5, f"- {clean_pdf_text(m[0])} ({m[1]})", ln=True)

    pdf.set_xy(10, y_radar + 95)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, "C. PREVIEW KEDISIPLINAN & EKSTRAKURIKULER", ln=True)
    y_pie = pdf.get_y()
    
    if absensi_synced and 'Sakit' in siswa_data:
        sakit, izin, alpa = pd.to_numeric(siswa_data.get('Sakit', 0)), pd.to_numeric(siswa_data.get('Izin', 0)), pd.to_numeric(siswa_data.get('Alpa', 0))
        hadir = max(0, 120 - (sakit + izin + alpa))
        df_absen = pd.DataFrame({'Kat': ['Hadir', 'Sakit', 'Izin', 'Alpa'], 'Jml': [hadir, sakit, izin, alpa]})
        fig_pie = px.pie(df_absen, values='Jml', names='Kat', hole=0.4, color='Kat', color_discrete_map={'Hadir':'#2ca02c', 'Sakit':'#ff7f0e', 'Izin':'#1f77b4', 'Alpa':'#d62728'})
        fig_pie.update_layout(margin=dict(l=0, r=0, t=0, b=0), showlegend=True)
        
        tmp_pie = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp_pie_name = tmp_pie.name
        tmp_pie.close()
        try:
            fig_pie.write_image(tmp_pie_name, width=350, height=250)
            pdf.image(tmp_pie_name, x=10, y=y_pie, w=80)
        finally:
            if os.path.exists(tmp_pie_name): os.remove(tmp_pie_name)
            
        pdf.set_xy(110, y_pie + 10)
        pdf.set_font("Helvetica", "B", 10); pdf.cell(0, 5, "Partisipasi Ekstrakurikuler:", ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_x(110); pdf.cell(0, 5, f"- Pramuka  : {siswa_data.get('Ekskul_Pramuka', '-')}", ln=True)
        pdf.set_x(110); pdf.cell(0, 5, f"- Kesenian : {siswa_data.get('Ekskul_Kesenian', '-')}", ln=True)
        pdf.set_x(110); pdf.cell(0, 5, f"- Olahraga : {siswa_data.get('Ekskul_Olahraga', '-')}", ln=True)
    else:
        pdf.set_font("Helvetica", "I", 10); pdf.set_x(15); pdf.cell(0, 10, "(Data absensi, ekskul, dan catatan BK belum disinkronkan)", ln=True)

    pdf.add_page(); pdf.set_left_margin(10); pdf.set_right_margin(10); pdf.set_x(10)
    
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 8, "D. RINCIAN NILAI MATA PELAJARAN", ln=True)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(100, 7, " Mata Pelajaran", 1, 0, 'L'); pdf.cell(45, 7, "Nilai Rata-rata", 1, 1, 'C')
    
    pdf.set_font("Helvetica", "", 10)
    for m in mapel_list:
        pdf.cell(100, 6, f" {clean_pdf_text(m)}", 1, 0, 'L'); pdf.cell(45, 6, str(siswa_data.get(m, 0)), 1, 1, 'C')
    pdf.ln(6)
    
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 8, "E. ANALISIS PERILAKU, KEHADIRAN & EKSTRAKURIKULER", ln=True)
    pdf.set_font("Helvetica", "", 10)
    
    if absensi_synced and 'Sakit' in siswa_data:
        sakit = int(pd.to_numeric(siswa_data.get('Sakit', 0), errors='coerce') or 0)
        izin = int(pd.to_numeric(siswa_data.get('Izin', 0), errors='coerce') or 0)
        alpa = int(pd.to_numeric(siswa_data.get('Alpa', 0), errors='coerce') or 0)
        
        deskripsi_absen = f"Berdasarkan rekapitulasi kehadiran semester ini, {nama_clean} mencatatkan "
        catatan_tidak_hadir = []
        if sakit > 0: catatan_tidak_hadir.append(f"{sakit} hari Sakit")
        if izin > 0: catatan_tidak_hadir.append(f"{izin} hari Izin")
        if alpa > 0: catatan_tidak_hadir.append(f"{alpa} hari Alpa (Tanpa Keterangan)")
        
        if len(catatan_tidak_hadir) == 0: deskripsi_absen += "tingkat kehadiran sempurna (100% Hadir) tanpa riwayat absen."
        else:
            deskripsi_absen += ", ".join(catatan_tidak_hadir) + ". "
            if alpa > 0: deskripsi_absen += "Adanya catatan Alpa memerlukan perhatian khusus agar tidak mengganggu ketertinggalan materi akademiknya."
            else: deskripsi_absen += "Tingkat kehadiran siswa masih wajar dan mendukung kelancaran proses kegiatan belajar."
                
        prak, seni, olr = siswa_data.get('Ekskul_Pramuka', '-'), siswa_data.get('Ekskul_Kesenian', '-'), siswa_data.get('Ekskul_Olahraga', '-')
        deskripsi_ekskul = f"Dalam pengembangan diri non-akademik, capaian predikat: Pramuka [{prak}], Kesenian [{seni}], dan Olahraga [{olr}]."
        if 'A' in [prak, seni, olr]: deskripsi_ekskul += " Siswa menunjukkan minat aktif dan potensi kepemimpinan yang baik."
            
        ctt_bk = str(siswa_data.get('Catatan BK', '')).strip()
        if not ctt_bk or ctt_bk == 'nan' or ctt_bk == '-': deskripsi_bk = "Catatan BK: Tidak terdapat catatan pelanggaran disiplin."
        else: deskripsi_bk = f"Catatan Bimbingan Konseling (BK): {ctt_bk}"
            
        pdf.multi_cell(190, 5, clean_pdf_text(deskripsi_absen), align="J"); pdf.ln(2)
        pdf.multi_cell(190, 5, clean_pdf_text(deskripsi_ekskul), align="J"); pdf.ln(2)
        pdf.set_font("Helvetica", "I", 10); pdf.multi_cell(190, 5, clean_pdf_text(deskripsi_bk), align="J"); pdf.set_font("Helvetica", "", 10)
    else:
        pdf.set_font("Helvetica", "I", 10); pdf.cell(0, 5, "Analisis kehadiran, ekskul, dan BK tidak dapat disajikan karena data belum disinkronkan.", ln=True)
    pdf.ln(4)
    
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 8, "F. INTERPRETASI WALI KELAS & SARAN AKADEMIK", ln=True)
    if len(mapel_urut) >= 2:
        top1, top2 = clean_pdf_text(mapel_urut[0][0]), clean_pdf_text(mapel_urut[1][0])
        bot1, bot2 = clean_pdf_text(mapel_urut[-1][0]), clean_pdf_text(mapel_urut[-2][0])
    else: top1, top2, bot1, bot2 = "-", "-", "-", "-"
    
    teks_grafik = f"Berdasarkan profil kompetensi akademik, {nama_clean} dominan pada mata pelajaran {top1} dan {top2}. Sebaliknya, perlu adanya perhatian khusus pada pelajaran {bot1} dan {bot2} yang menjadi tantangannya."
    pdf.set_x(10); pdf.multi_cell(190, 5, teks_grafik, align="J"); pdf.ln(3)
    
    if pd.to_numeric(rata_rata, errors='coerce') >= 85: saran = f"Saran Wali Kelas: Pertahankan prestasimu! Gunakan metode belajar kelompok untuk menaklukkan pelajaran {bot1}."
    elif pd.to_numeric(rata_rata, errors='coerce') >= 75: saran = f"Saran Wali Kelas: Progres belajarmu sudah stabil. Sisihkan waktu ekstra setiap hari untuk mengulang materi {bot1}."
    else: saran = f"Saran Wali Kelas: Saatnya bangkit! Sangat disarankan mengikuti bimbingan tambahan (remedial) untuk pelajaran {bot1}."
                 
    pdf.set_x(10); pdf.set_font("Helvetica", "BI", 10); pdf.multi_cell(190, 5, saran, align="J")
    return bytes(pdf.output())

# --- FUNGSI LOAD DATA UTAMA & DETEKSI KELAS ---
@st.cache_data
def process_leger(file_buffer, file_name):
    if hasattr(file_buffer, 'seek'): file_buffer.seek(0)
    try:
        if file_name.endswith('.csv'): df_raw = pd.read_csv(file_buffer, header=None, dtype=str).fillna('')
        elif file_name.endswith('.ods'): df_raw = pd.read_excel(file_buffer, engine="odf", header=None, dtype=str).fillna('')
        else: df_raw = pd.read_excel(file_buffer, header=None, dtype=str).fillna('')
            
        kelas_name = "Kelas Umum"
        for i in range(min(15, len(df_raw))):
            row_str = ' '.join(df_raw.iloc[i].astype(str).tolist()).upper()
            if 'KELAS' in row_str and ':' in row_str:
                try:
                    extracted = row_str.split(':')[1].replace('NAN', '').replace(',', '').strip()
                    if extracted: kelas_name = extracted; break
                except: pass
                    
        header_idx = 0
        for i in range(min(30, len(df_raw))):
            row_str = [str(cell).upper() for cell in df_raw.iloc[i].tolist()]
            if any('NAMA' in cell for cell in row_str):
                header_idx = i; break
                
        smt_row_idx = None
        for i in range(header_idx, min(header_idx + 15, len(df_raw))):
            row_str = [str(x).lower() for x in df_raw.iloc[i].tolist()]
            if any('smt1' in x or 'smt 1' in x for x in row_str):
                smt_row_idx = i; break
                
        if smt_row_idx is not None:
            header_row_vals = [str(val).upper() for val in df_raw.iloc[header_idx].tolist()]
            col_nama, col_nis, col_nisn = 1, 3, 2 
            for idx, val in enumerate(header_row_vals):
                if 'NAMA' in val: col_nama = idx
                elif 'NISN' in val: col_nisn = idx
                elif 'NIS' in val and 'NISN' not in val: col_nis = idx
            
            data_start = smt_row_idx + 1
            df_data = df_raw.iloc[data_start:].copy()
            df_data = df_data[df_data[col_nama].notna() & (df_data[col_nama].astype(str).str.strip() != '')]
            df_data = df_data[~df_data[col_nama].astype(str).str.upper().str.contains('NAMA SISWA|TOTAL|AVERAGE|RELA|NAN', na=False)]
            
            df_clean = pd.DataFrame()
            df_clean['Nama'] = df_data[col_nama].astype(str).str.strip()
            df_clean['NISN'] = df_data[col_nisn].astype(str).str.replace(r'\.0$', '', regex=True).str.zfill(10)
            df_clean['NIS'] = df_data[col_nis].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
            subject_row_idx = smt_row_idx - 1
            while subject_row_idx > header_idx:
                sampel = [str(x).lower() for x in df_raw.iloc[subject_row_idx, col_nis+1:col_nis+10].tolist()]
                if any(x not in ['nan', '', 'none', 'unnamed'] for x in sampel): break
                subject_row_idx -= 1
                
            subjects = [str(x).strip() for x in df_raw.iloc[subject_row_idx].tolist()]
            smt_labels = [str(x).lower().strip() for x in df_raw.iloc[smt_row_idx].tolist()]
            mapel_list, current_subj, mapel_cols = [], None, {}
            
            for c in range(col_nis + 1, len(df_raw.columns)):
                sub_val = subjects[c]
                if sub_val.lower() not in ['nan', '', 'none', 'unnamed']:
                    if 'agama' in sub_val.lower() or 'pai' in sub_val.lower(): current_subj = 'PAI'
                    elif 'pancasila' in sub_val.lower() and 'profil' not in sub_val.lower(): current_subj = 'PPKn'
                    elif 'bahasa indonesia' in sub_val.lower(): current_subj = 'B.IND'
                    elif 'bahasa inggris' in sub_val.lower(): current_subj = 'B.ING'
                    elif 'matematika' in sub_val.lower(): current_subj = 'MTK'
                    elif 'alam' in sub_val.lower() or 'ipa' in sub_val.lower(): current_subj = 'IPA'
                    elif 'sosial' in sub_val.lower() or 'ips' in sub_val.lower(): current_subj = 'IPS'
                    elif 'jasmani' in sub_val.lower() or 'olahraga' in sub_val.lower(): current_subj = 'PJOK'
                    elif 'seni' in sub_val.lower() or 'budaya' in sub_val.lower(): current_subj = 'SENI'
                    elif 'informatika' in sub_val.lower(): current_subj = 'INF'
                    elif 'sunda' in sub_val.lower() or 'daerah' in sub_val.lower(): current_subj = 'B.SUN'
                    elif 'project' in sub_val.lower() or 'p5' in sub_val.lower() or 'profil' in sub_val.lower(): current_subj = None 
                    else: current_subj = sub_val[:12]
                    
                    if current_subj and current_subj not in mapel_cols:
                        mapel_cols[current_subj] = {'semesters': [], 'rerata': None, 'smt1': None, 'smt3': None}
                
                if current_subj:
                    lbl = smt_labels[c]
                    if 'smt' in lbl or 'sem' in lbl:
                        mapel_cols[current_subj]['semesters'].append(c)
                        if '1' in lbl: mapel_cols[current_subj]['smt1'] = c
                        if '3' in lbl: mapel_cols[current_subj]['smt3'] = c
                    elif 'rerata' in lbl or 'rata' in lbl: mapel_cols[current_subj]['rerata'] = c
                        
            tren_list_smt3, tren_list_smt1 = [], []
            for m, info in mapel_cols.items():
                if info['rerata'] is not None and str(df_data[info['rerata']].dropna().max()).strip() != '':
                    df_clean[m] = pd.to_numeric(df_data[info['rerata']], errors='coerce').round(2)
                    mapel_list.append(m)
                elif info['semesters']:
                    df_clean[m] = pd.to_numeric(df_data[info['semesters']].astype(float).mean(axis=1), errors='coerce').round(2)
                    mapel_list.append(m)
                    
                if info['smt1'] is not None and info['smt3'] is not None:
                    s1_val = pd.to_numeric(df_data[info['smt1']], errors='coerce')
                    s3_val = pd.to_numeric(df_data[info['smt3']], errors='coerce')
                    tren_list_smt1.append(s1_val)
                    tren_list_smt3.append(s3_val)
                    df_clean[f'Tren_{m}'] = (s3_val.fillna(0) - s1_val.fillna(0)).round(2)
                else: df_clean[f'Tren_{m}'] = 0.0
            
            mapel_list = [m for m in mapel_list if df_clean[m].notna().any()]
            df_clean['Total Nilai'] = df_clean[mapel_list].sum(axis=1, skipna=True)
            df_clean['Rata-rata'] = df_clean[mapel_list].mean(axis=1, skipna=True).round(2)
            
            if tren_list_smt1 and tren_list_smt3:
                df_clean['Tren_Belajar'] = (pd.concat(tren_list_smt3, axis=1).mean(axis=1) - pd.concat(tren_list_smt1, axis=1).mean(axis=1)).round(2)
            else: df_clean['Tren_Belajar'] = 0.0
            df = df_clean
        else:
            df = df_raw.copy()
            df.columns = df.iloc[header_idx].astype(str).str.strip()
            df = df.iloc[header_idx+1:].copy()
            df = df.loc[:, ~df.columns.duplicated()].copy()
            df['Nama'] = df.iloc[:, 1].astype(str).str.strip()
            df['Total Nilai'] = 0
            df['Rata-rata'] = 0
            df['Tren_Belajar'] = 0
            mapel_list = []

        def get_badge(row):
            if row['Rata-rata'] >= 85: return "🌟 High Achiever"
            elif row['Rata-rata'] < 75: return "🆘 Fokus Perbaikan"
            else: return "✅ Stabil"
                
        df['Badge'] = df.apply(get_badge, axis=1)
        df = df.sort_values(by='Total Nilai', ascending=False).reset_index(drop=True)
        df.index = df.index + 1
        return df, mapel_list, kelas_name
        
    except Exception as e:
        st.error(f"❌ Gagal memproses file Leger. Detail error: {str(e)}")
        st.stop()

# Eksekusi Pemrosesan File
df_utama, mapel_list, kelas_name = process_leger(file_leger, file_leger.name)

if 'last_leger_name' not in st.session_state or st.session_state.last_leger_name != file_leger.name:
    st.session_state.df_lengkap = df_utama.copy()
    st.session_state.absensi_synced = False
    st.session_state.last_leger_name = file_leger.name
    st.session_state.nama_kelas = kelas_name
    try: st.rerun() 
    except AttributeError: st.experimental_rerun()

df = st.session_state.df_lengkap.copy()
if 'Total Nilai' in df.columns:
    df = df.sort_values(by='Total Nilai', ascending=False).reset_index(drop=True)
    df.index = df.index + 1

# --- TABS NAVIGASI UTAMA ---
t1, t2, t3, t4, t5, t6 = st.tabs([
    "👤 Profil Siswa", 
    "⚖️ Komparasi Head-to-Head", 
    "📚 Evaluasi Mapel", 
    "🤖 AI Klasterisasi Siswa", 
    "🔄 Sync Data", 
    "📊 Database Kelas"
])

# ==============================================================================
# TAB 1: PROFIL & PERSONALISASI (UI SEGMENTED CLEAN)
# ==============================================================================
with t1:
    pilihan_nama = st.selectbox("Cari Profil Siswa:", df['Nama'].tolist())
    
    if pilihan_nama:
        siswa_idx = df[df['Nama'] == pilihan_nama].index[0]
        siswa_data = df[df['Nama'] == pilihan_nama].iloc[0]
        
        st.markdown(f"## 🎓 {pilihan_nama} | {siswa_data.get('Badge', '-')}")
        st.write(f"**NISN:** {siswa_data.get('NISN', '-')} | **NIS:** {siswa_data.get('NIS', '-')} | **Peringkat:** {siswa_idx} dari {len(df)}")
        
        dict_mapel = {m: pd.to_numeric(siswa_data.get(m, 0), errors='coerce') for m in mapel_list}
        mapel_urut = sorted(dict_mapel.items(), key=lambda x: x[1] if pd.notnull(x[1]) else 0, reverse=True)
        
        bonus_pramuka = get_ekskul_bonus(siswa_data.get('Ekskul_Pramuka', '-'))
        bonus_seni = get_ekskul_bonus(siswa_data.get('Ekskul_Kesenian', '-'))
        bonus_olahraga = get_ekskul_bonus(siswa_data.get('Ekskul_Olahraga', '-'))
        
        nilai_mtk = pd.to_numeric(siswa_data.get('MTK', 0), errors='coerce') or 0
        nilai_ipa = pd.to_numeric(siswa_data.get('IPA', 0), errors='coerce') or 0
        nilai_ips = pd.to_numeric(siswa_data.get('IPS', 0), errors='coerce') or 0
        nilai_bind = pd.to_numeric(siswa_data.get('B.IND', 0), errors='coerce') or 0
        nilai_bing = pd.to_numeric(siswa_data.get('B.ING', 0), errors='coerce') or 0
        nilai_inf = pd.to_numeric(siswa_data.get('INF', 0), errors='coerce') or 0
        nilai_seni = pd.to_numeric(siswa_data.get('SENI', 0), errors='coerce') or 0
        
        skor_penjurusan = {
            "SMA: MIPA / Sains": ((nilai_mtk + nilai_ipa) / 2) + (bonus_pramuka * 0.5),
            "SMA: Soshum / IPS": ((nilai_ips + nilai_bind) / 2) + (bonus_pramuka * 0.5),
            "SMA: Ilmu Bahasa & Budaya": ((nilai_bind + nilai_bing) / 2) + (bonus_seni * 0.3),
            "SMK: Rekayasa Teknologi & IT": ((nilai_inf + nilai_mtk + nilai_bing) / 3) + (bonus_olahraga * 0.2),
            "SMK: Seni, Kreatif & Multimedia": ((nilai_seni + nilai_inf) / 2) + bonus_seni
        }
        rek_urut = sorted(skor_penjurusan.items(), key=lambda x: x[1], reverse=True)

        st.markdown("<br>", unsafe_allow_html=True)
        
        # 🌟 PILIHAN MENU SUB-ANALISIS (UI CLEAN)
        menu_terpilih = st.radio(
            "Visualisasi Aspek Profiling Siswa:",
            ["📝 Resume & PDF", "📊 Analisis Akademik", "⏰ Kedisiplinan", "🏕️ Ekskul & Bakat", "🛡️ Catatan BK"],
            horizontal=True,
            key="sub_menu_siswa",
            label_visibility="collapsed"
        )
        st.markdown("---")
        
        # --- SUB-MENU 1: RESUME ---
        if menu_terpilih == "📝 Resume & PDF":
            st.subheader("📋 Ringkasan Profil & Output Holistik")
            m1, m2, m3 = st.columns(3)
            m1.metric("Rata-rata Akademik", f"{siswa_data.get('Rata-rata', 0)}")
            m2.metric("Total Skor", f"{siswa_data.get('Total Nilai', 0):.0f}")
            m3.metric("Tren Belajar", f"{siswa_data.get('Tren_Belajar', 0)} Poin", delta=float(siswa_data.get('Tren_Belajar', 0)))
            
            st.markdown("##### 📌 Kesimpulan Wali Kelas:")
            rata_rata = pd.to_numeric(siswa_data.get('Rata-rata', 0), errors='coerce')
            if rata_rata >= 85: st.success("✨ Pertahankan prestasimu yang luar biasa ini! Nilai akademik yang prima menunjukkan kesiapan tinggi untuk program pengayaan kompetensi lanjutan.")
            elif rata_rata >= 75: st.info("📈 Progres belajarmu sudah baik dan stabil. Berikan perhatian konsisten serta sisihkan waktu ekstra khusus untuk materi tantangan.")
            else: st.warning("⚠️ Saatnya bangkit! Mari perbaiki fokus belajar dengan membatasi distraksi di luar jam belajar. Sangat disarankan mengikuti program bimbingan terarah.")
                
            st.markdown("### 📥 Ekspor Laporan Resmi")
            try:
                pdf_data = generate_student_pdf(siswa_data, mapel_list, siswa_idx, len(df), rek_urut, mapel_urut, st.session_state.nama_kelas, st.session_state.absensi_synced)
                st.download_button(label=f"📄 Download Laporan PDF - {pilihan_nama}", data=pdf_data, file_name=f"Laporan_Dashboard_{pilihan_nama.replace(' ', '_')}.pdf", mime="application/pdf")
            except Exception as pdf_err: st.error(f"Gagal memproses file PDF: {pdf_err}")

        # --- SUB-MENU 2: ANALISIS AKADEMIK (DENGAN TREN AI) ---
        elif menu_terpilih == "📊 Analisis Akademik":
            col_radar, col_jalur = st.columns([1.3, 1])
            with col_radar:
                st.subheader("🎯 Peta Kompetensi Mata Pelajaran")
                if mapel_list:
                    categories = mapel_list
                    values = [siswa_data.get(m, 0) for m in mapel_list]
                    fig = go.Figure()
                    fig.add_trace(go.Scatterpolar(r=values, theta=categories, fill='toself', name=pilihan_nama))
                    fig.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])), margin=dict(l=30, r=30, t=30, b=30))
                    st.plotly_chart(fig, use_container_width=True)
                    
            with col_jalur:
                st.subheader("🛤️ Prediksi Penjurusan (AI Hybrid)")
                if rek_urut and len(rek_urut) >= 2:
                    st.info(f"**Rekomendasi Utama:**\n\n🥇 {rek_urut[0][0]}\n\n🥈 {rek_urut[1][0]}")
                if len(mapel_urut) >= 3:
                    st.success("🌟 **Kekuatan Utama Belajar:**\n" + "\n".join([f"- {m[0]} ({m[1]})" for m in mapel_urut[:3]]))
                    st.error("📉 **Area Fokus Perbaikan:**\n" + "\n".join([f"- {m[0]} ({m[1]})" for m in mapel_urut[-3:]]))

            st.markdown("---")
            st.subheader("📈 Analisis Perkembangan Kompetensi (Riwayat Tren)")
            if mapel_list:
                mapel_naik, mapel_turun = [], []
                for mapel in mapel_list:
                    kolom_tren = f'Tren_{mapel}'
                    if kolom_tren in siswa_data and pd.notna(siswa_data[kolom_tren]):
                        perubahan = float(siswa_data[kolom_tren])
                        if perubahan > 0: mapel_naik.append(f"{mapel} (+{perubahan})")
                        elif perubahan < 0: mapel_turun.append(f"{mapel} ({perubahan})")
                
                col_ringkasan_naik, col_ringkasan_turun = st.columns(2)
                with col_ringkasan_naik:
                    if mapel_naik: st.success(f"🔼 **Mengalami Peningkatan di:** \n{', '.join(mapel_naik)}")
                    else: st.write("✨ *Belum ada mapel dengan peningkatan signifikan.*")
                with col_ringkasan_turun:
                    if mapel_turun: st.error(f"⚠️ **Butuh Perhatian (Menurun) di:** \n{', '.join(mapel_turun)}")
                    else: st.info("✅ *Hebat! Tidak ada mata pelajaran yang menurun.*")
                
                st.markdown(" ") 
                mapel_terpilih = st.selectbox("🎯 Pilih Mata Pelajaran untuk melihat visualisasi detail:", mapel_list, key="dropdown_tren_mapel")
                kolom_tren_terpilih = f'Tren_{mapel_terpilih}'
                if kolom_tren_terpilih in siswa_data and pd.notna(siswa_data[kolom_tren_terpilih]):
                    perubahan_terpilih = float(siswa_data[kolom_tren_terpilih])
                    st.markdown(f"#### Detail Tren: **{mapel_terpilih}**")
                    col_info_tren, col_grafik_tren = st.columns([1, 2])
                    
                    with col_info_tren:
                        if perubahan_terpilih > 0: st.success(f"📈 **Status: Positif** \nSiswa berkembang **+{perubahan_terpilih} poin** dari sebelumnya.")
                        elif perubahan_terpilih < 0: st.error(f"📉 **Status: Kritis** \nSiswa merosot **{perubahan_terpilih} poin**. Perlu bimbingan.")
                        else: st.info(f"➖ **Status: Stabil** \nNilai stagnan (0 poin).")
                    
                    with col_grafik_tren:
                        df_titik = pd.DataFrame({'Periode': ['Periode Awal (Smt 1)', 'Periode Akhir (Smt 3)'], 'Progres Poin': [0, perubahan_terpilih]})
                        warna_garis = '#2ecc71' if perubahan_terpilih >= 0 else '#e74c3c'
                        fig_garis = px.line(df_titik, x='Periode', y='Progres Poin', markers=True, text='Progres Poin', labels={'Progres Poin': 'Selisih Nilai'})
                        fig_garis.update_traces(textposition="top center", line=dict(color=warna_garis, width=4), marker=dict(size=10))
                        fig_garis.update_layout(margin=dict(l=40, r=40, t=20, b=20), yaxis=dict(zeroline=True, zerolinecolor='grey'))
                        st.plotly_chart(fig_garis, use_container_width=True)

        # --- SUB-MENU 3: KEHADIRAN ---
        elif menu_terpilih == "⏰ Kedisiplinan":
            st.subheader("📋 Rekapitulasi Presensi & Indeks Disiplin")
            if st.session_state.absensi_synced:
                col_pie, col_details = st.columns([1.2, 1])
                with col_pie:
                    sakit, izin, alpa = siswa_data.get('Sakit', 0), siswa_data.get('Izin', 0), siswa_data.get('Alpa', 0)
                    hadir = max(0, 120 - (sakit + izin + alpa))
                    df_absen = pd.DataFrame({'Kategori': ['Hadir', 'Sakit', 'Izin', 'Alpa'], 'Jumlah': [hadir, sakit, izin, alpa]})
                    fig_pie = px.pie(df_absen, values='Jumlah', names='Kategori', hole=0.4, color='Kategori', color_discrete_map={'Hadir':'#2ca02c', 'Sakit':'#ff7f0e', 'Izin':'#1f77b4', 'Alpa':'#d62728'})
                    st.plotly_chart(fig_pie, use_container_width=True)
                with col_details:
                    st.markdown("##### 📊 Metrik Presensi Aktual:")
                    st.write(f"🤒 **Sakit:** {siswa_data.get('Sakit', 0)} Hari")
                    st.write(f"📨 **Izin:** {siswa_data.get('Izin', 0)} Hari")
                    st.write(f"❌ **Alpa (Tanpa Keterangan):** {siswa_data.get('Alpa', 0)} Hari")
                    if 'Skor_Disiplin' in siswa_data: st.metric("Skor Integritas Disiplin", f"{siswa_data.get('Skor_Disiplin', 0)}")
            else: st.warning("⚠️ Data kehadiran belum disinkronkan. Sila unggah file presensi sekunder di tab **'Sync Data'**.")

        # --- SUB-MENU 4: EKSKUL ---
        elif menu_terpilih == "🏕️ Ekskul & Bakat":
            st.subheader("🏅 Rekapitulasi Ekstrakurikuler & Pengembangan Diri")
            if st.session_state.absensi_synced:
                c1, c2, c3 = st.columns(3)
                c1.info(f"🏕️ **Pramuka (Wajib):** \n\nPredikat: **{siswa_data.get('Ekskul_Pramuka', '-')}**")
                c2.info(f"🎨 **Kesenian & Budaya:** \n\nPredikat: **{siswa_data.get('Ekskul_Kesenian', '-')}**")
                c3.info(f"⚽ **Minat Olahraga:** \n\nPredikat: **{siswa_data.get('Ekskul_Olahraga', '-')}**")
            else: st.warning("⚠️ Data penilaian ekstrakurikuler belum dimuat. Selesaikan sinkronisasi data sekunder terlebih dahulu.")

        # --- SUB-MENU 5: BK ---
        elif menu_terpilih == "🛡️ Catatan BK":
            st.subheader("🛡️ Log Kasus & Bimbingan Konseling")
            if 'Catatan BK' in siswa_data and str(siswa_data.get('Catatan BK', '')).strip() not in ['', '-', 'nan']:
                st.error(f"⚠️ **Ditemukan Catatan Khusus/Pelanggaran:**")
                st.markdown(f"> {siswa_data.get('Catatan BK')}")
            else: st.success("✅ **Siswa Bersih & Adaptif** \nTidak ada catatan pelanggaran tata tertib atau rekam kasus negatif dari tim Bimbingan Konseling (BK).")

# ==============================================================================
# TAB 2: KOMPARASI HEAD-TO-HEAD (DENGAN TABEL & AI PREDICTIVE INSIGHTS)
# ==============================================================================
with t2:
    st.header("⚖️ Komparasi Head-to-Head Siswa")
    col_s1, col_s2 = st.columns(2)
    with col_s1: siswa1 = st.selectbox("Siswa 1:", df['Nama'].tolist(), key="s1")
    with col_s2: siswa2 = st.selectbox("Siswa 2:", df['Nama'].tolist(), key="s2")
    
    if siswa1 and siswa2 and mapel_list:
        d1, d2 = df[df['Nama'] == siswa1].iloc[0], df[df['Nama'] == siswa2].iloc[0]
        categories = mapel_list
        fig_vs = go.Figure()
        fig_vs.add_trace(go.Scatterpolar(r=[d1.get(m,0) for m in mapel_list], theta=categories, fill='toself', name=siswa1))
        fig_vs.add_trace(go.Scatterpolar(r=[d2.get(m,0) for m in mapel_list], theta=categories, fill='toself', name=siswa2))
        fig_vs.update_layout(polar=dict(radialaxis=dict(visible=True, range=[0, 100])), margin=dict(l=20, r=20, t=20, b=20))
        st.plotly_chart(fig_vs, use_container_width=True)
        
        st.markdown("---")
        st.subheader("📊 Ringkasan Perbandingan Akademik")
        rank1, rank2 = df[df['Nama'] == siswa1].index[0], df[df['Nama'] == siswa2].index[0]
        avg1, avg2 = d1.get('Rata-rata', 0), d2.get('Rata-rata', 0)
        tot1, tot2 = d1.get('Total Nilai', 0), d2.get('Total Nilai', 0)
        tren1, tren2 = d1.get('Tren_Belajar', 0), d2.get('Tren_Belajar', 0)
        
        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Peringkat", f"#{rank1}", help=siswa1)
        mc2.metric("Peringkat", f"#{rank2}", help=siswa2)
        mc3.metric(f"Rata-rata {siswa1.split()[0]}", f"{avg1}")
        mc4.metric(f"Rata-rata {siswa2.split()[0]}", f"{avg2}", delta=f"{avg2-avg1:+.2f} vs {siswa1.split()[0]}")
        
        # --- TABEL KOMPARATIF BERWARNA ---
        st.markdown("#### 📋 Tabel Rincian Keunggulan per Mata Pelajaran")
        col_s1_name, col_s2_name = f"{siswa1} (S1)", f"{siswa2} (S2)"
        selisih_col = f"Selisih ({siswa2.split()[0]} - {siswa1.split()[0]})"
        
        rows = []
        s1_strong, s2_strong = [], []
        for m in mapel_list:
            n1 = pd.to_numeric(d1.get(m, 0), errors='coerce') or 0
            n2 = pd.to_numeric(d2.get(m, 0), errors='coerce') or 0
            selisih = round(n2 - n1, 2)
            if n1 > n2: 
                unggul = f"✅ {siswa1.split()[0]}"
                s1_strong.append(m)
            elif n2 > n1: 
                unggul = f"✅ {siswa2.split()[0]}"
                s2_strong.append(m)
            else: unggul = "🤝 Seri"
            rows.append({"Mata Pelajaran": m, col_s1_name: n1, col_s2_name: n2, selisih_col: selisih, "Keunggulan": unggul})
        
        df_compare = pd.DataFrame(rows)
        df_display = df_compare.copy()
        df_display[col_s1_name] = df_display[col_s1_name].apply(lambda x: f"{x:.2f}")
        df_display[col_s2_name] = df_display[col_s2_name].apply(lambda x: f"{x:.2f}")
        df_display[selisih_col] = df_display[selisih_col].apply(lambda x: f"{x:+.2f}")
        
        def style_row(row):
            styles = [''] * len(row)
            idx = list(row.index).index(selisih_col)
            try:
                v = float(row[selisih_col])
                if v > 0: styles[idx] = 'background-color: #d4edda; color: #155724; font-weight: bold'
                elif v < 0: styles[idx] = 'background-color: #f8d7da; color: #721c24; font-weight: bold'
                else: styles[idx] = 'background-color: #e2e3e5; color: #383d41'
            except: pass
            return styles
        
        styled_df = df_display.style.apply(style_row, axis=1)\
            .set_properties(**{'text-align': 'center'}, subset=[col_s1_name, col_s2_name, selisih_col, "Keunggulan"])\
            .set_table_styles([{'selector': 'thead th', 'props': [('background-color', '#343a40'), ('color', 'white'), ('text-align', 'center'), ('font-weight', 'bold')]}])
        
        st.dataframe(styled_df, use_container_width=True, hide_index=True)
        
        # --- DESKRIPSI AI PREDIKTIF KOMPARASI HOLISTIK ---
        st.markdown("---")
        st.subheader("🤖 AI Predictive Insights - Analisis Komparatif & Proyeksi Karir")
        
        leader = siswa1 if avg1 > avg2 else siswa2
        follower = siswa2 if avg1 > avg2 else siswa1
        margin = abs(avg1 - avg2)
        
        def compute_career_vector(row_s):
            bp = get_ekskul_bonus(row_s.get('Ekskul_Pramuka', '-'))
            bs = get_ekskul_bonus(row_s.get('Ekskul_Kesenian', '-'))
            bo = get_ekskul_bonus(row_s.get('Ekskul_Olahraga', '-'))
            v_mtk, v_ipa = pd.to_numeric(row_s.get('MTK', 0), errors='coerce') or 0, pd.to_numeric(row_s.get('IPA', 0), errors='coerce') or 0
            v_ips, v_bind = pd.to_numeric(row_s.get('IPS', 0), errors='coerce') or 0, pd.to_numeric(row_s.get('B.IND', 0), errors='coerce') or 0
            v_bing, v_inf = pd.to_numeric(row_s.get('B.ING', 0), errors='coerce') or 0, pd.to_numeric(row_s.get('INF', 0), errors='coerce') or 0
            v_seni = pd.to_numeric(row_s.get('SENI', 0), errors='coerce') or 0
            return {
                "Sains / MIPA": ((v_mtk + v_ipa) / 2) + (bp * 0.5),
                "Soshum / Sosial": ((v_ips + v_bind) / 2) + (bp * 0.5),
                "Bahasa & Budaya": ((v_bind + v_bing) / 2) + (bs * 0.3),
                "Teknologi & IT": ((v_inf + v_mtk + v_bing) / 3) + (bo * 0.2),
                "Seni Kreatif": ((v_seni + v_inf) / 2) + bs
            }
            
        vec1 = sorted(compute_career_vector(d1).items(), key=lambda x: x[1], reverse=True)[0][0]
        vec2 = sorted(compute_career_vector(d2).items(), key=lambda x: x[1], reverse=True)[0][0]
        
        narrative = f"### 💡 Hasil Profiling Sistem Cerdas:\n"
        if margin == 0: narrative += f"- **Dominasi Akademik:** Kedua siswa berada pada keseimbangan akademis makro yang setara secara keseluruhan.\n"
        else: narrative += f"- **Dominasi Akademik:** **{leader}** memegang keunggulan makro kelas dengan margin selisih **+{margin:.2f} poin rata-rata** di atas **{follower}**.\n"
            
        narrative += f"- **Spesialisasi Subjek:** **{siswa1.split()[0]}** unggul dalam penguasaan {', '.join(s1_strong[:3]) if s1_strong else 'umum'}, sedangkan **{siswa2.split()[0]}** menunjukkan ketajaman pada rumpun {', '.join(s2_strong[:3]) if s2_strong else 'umum'}.\n\n"
        
        narrative += f"### 📈 Proyeksi Trajektori & Kecepatan Belajar:\n"
        if tren1 > 0 and tren2 > 0:
            narrative += f"- Keduanya mencatat pertumbuhan positif (+{tren1} vs +{tren2} Poin). Pola ini mendeteksi lingkungan belajar yang kondusif. Akselerasi pemahaman **{siswa1 if tren1 > tren2 else siswa2}** diproyeksikan berkembang lebih tangkas secara statistik.\n"
        elif tren1 > 0 and tren2 <= 0:
            narrative += f"- ⚠️ Peringatan Sistem: Indikasi penurunan tren belajar terdeteksi pada **{siswa2}** ({tren2} poin). Sebaliknya, **{siswa1}** berada dalam fase ekspansi daya serap aktif (+{tren1} poin).\n"
        elif tren2 > 0 and tren1 <= 0:
            narrative += f"- ⚠️ Peringatan Sistem: Indikasi penurunan tren belajar terdeteksi pada **{siswa1}** ({tren1} poin). Sebaliknya, **{siswa2}** berada dalam fase ekspansi daya serap aktif (+{tren2} poin).\n"
        else:
            narrative += f"- ⚠️ Peringatan Bersama: Keduanya mengalami degradasi tren akumulasi nilai. Direkomendasikan evaluasi pola istirahat atau beban tugas di rumah.\n\n"
            
        narrative += f"### 🛤️ Sinergi Penjurusan Karir Masa Depan:\n"
        if vec1 == vec2:
            narrative += f"- **Pola Sinkronisasi Karir:** Sistem mendeteksi kedua siswa memiliki klaster minat masa depan yang identik (Cenderung pada bidang **{vec1}**). Keduanya sangat direkomendasikan menjadi partner belajar (*study buddy*).\n"
        else:
            narrative += f"- **Pola Divergensi Bakat:** Kedua siswa mengarah ke karir yang berbeda. **{siswa1.split()[0]}** optimal pada bidang **{vec1}**, sedangkan **{siswa2.split()[0]}** kuat pada **{vec2}**. Pendekatan bimbingan harus dipersonalisasi.\n"
            
        if st.session_state.absensi_synced:
            alpa1, alpa2 = int(pd.to_numeric(d1.get('Alpa', 0)) or 0), int(pd.to_numeric(d2.get('Alpa', 0)) or 0)
            if alpa1 > 0 or alpa2 > 0: narrative += f"\n⚠️ **Catatan Kontingensi:** Terdapat kerentanan kedisiplinan akibat catatan bolos (*Alpa*). Faktor non-kognitif ini wajib dibina agar tidak merusak potensi nilai di atas."
                
        st.info(narrative)
        
        st.markdown("---")
        st.subheader("📈 Visualisasi Perbandingan Metrik Lainnya")
        fig_bar_compare = go.Figure()
        fig_bar_compare.add_trace(go.Bar(name=siswa1, x=mapel_list, y=[pd.to_numeric(d1.get(m,0),errors='coerce') or 0 for m in mapel_list], marker_color='#636EFA'))
        fig_bar_compare.add_trace(go.Bar(name=siswa2, x=mapel_list, y=[pd.to_numeric(d2.get(m,0),errors='coerce') or 0 for m in mapel_list], marker_color='#EF553B'))
        fig_bar_compare.update_layout(barmode='group', xaxis_title="Mata Pelajaran", yaxis_title="Nilai", yaxis=dict(range=[0, 105]), margin=dict(l=20, r=20, t=30, b=20))
        st.plotly_chart(fig_bar_compare, use_container_width=True)

        if st.session_state.absensi_synced:
            st.markdown("### 🛡️ Komparasi Kehadiran & Catatan Perilaku")
            comp_col1, comp_col2 = st.columns(2)
            with comp_col1:
                st.markdown(f"**👤 {siswa1}**")
                st.write(f"🤒 Sakit: {d1.get('Sakit', 0)} | 📨 Izin: {d1.get('Izin', 0)} | ❌ Alpa: {d1.get('Alpa', 0)}")
                st.write(f"- Ekskul: Pramuka [{d1.get('Ekskul_Pramuka', '-')}], Kesenian [{d1.get('Ekskul_Kesenian', '-')}], Olahraga [{d1.get('Ekskul_Olahraga', '-')}]")
                st.info(f"Catatan BK: {d1.get('Catatan BK', 'Aman')}")
            with comp_col2:
                st.markdown(f"**👤 {siswa2}**")
                st.write(f"🤒 Sakit: {d2.get('Sakit', 0)} | 📨 Izin: {d2.get('Izin', 0)} | ❌ Alpa: {d2.get('Alpa', 0)}")
                st.write(f"- Ekskul: Pramuka [{d2.get('Ekskul_Pramuka', '-')}], Kesenian [{d2.get('Ekskul_Kesenian', '-')}], Olahraga [{d2.get('Ekskul_Olahraga', '-')}]")
                st.info(f"Catatan BK: {d2.get('Catatan BK', 'Aman')}")

# --- TAB 3: EVALUASI MATA PELAJARAN ---
with t3:
    st.header("📚 Evaluasi Kesulitan Mata Pelajaran")
    if mapel_list:
        avg_mapel = df[mapel_list].mean().round(2).sort_values()
        fig_bar = px.bar(x=avg_mapel.values, y=avg_mapel.index, orientation='h', text=avg_mapel.values,
                         labels={'x': 'Rata-rata Kelas', 'y': 'Mata Pelajaran'}, color=avg_mapel.values, color_continuous_scale='RdYlGn')
        st.plotly_chart(fig_bar, use_container_width=True)

# --- TAB 4: AI KLASTERISASI PREDIKTIF (K-MEANS) ---
with t4:
    st.header("🤖 AI Klasterisasi & Pemetaan Siswa Otomatis")
    st.write("Sistem Machine Learning (K-Means Clustering) ini secara otomatis membagi siswa ke dalam 3 kelompok utama berdasarkan metrik Rata-rata dan Tren Kecepatan Belajar mereka. Gunakan fitur ini untuk menentukan program pengayaan atau jadwal remedial massal.")
    
    if AI_AVAILABLE and not df.empty and 'Rata-rata' in df.columns:
        try:
            if 'Tren_Belajar' in df.columns and df['Tren_Belajar'].sum() != 0: X = df[['Rata-rata', 'Tren_Belajar']].fillna(0)
            else: X = df[['Rata-rata']].fillna(0)
                
            kmeans = KMeans(n_clusters=3, random_state=42)
            df['Cluster_ID'] = kmeans.fit_predict(X)
            
            cluster_means = df.groupby('Cluster_ID')['Rata-rata'].mean().sort_values()
            label_map = {
                cluster_means.index[0]: "🎯 Fokus Bimbingan (Bawah)",
                cluster_means.index[1]: "📈 Berkembang (Menengah)",
                cluster_means.index[2]: "🌟 Akselerasi (Atas)"
            }
            df['Kelompok AI'] = df['Cluster_ID'].map(label_map)
            
            if 'Tren_Belajar' in df.columns and df['Tren_Belajar'].sum() != 0:
                fig_ai = px.scatter(df, x="Rata-rata", y="Tren_Belajar", color="Kelompok AI", hover_name="Nama", size="Total Nilai",
                    color_discrete_map={"🌟 Akselerasi (Atas)": "#2ca02c", "📈 Berkembang (Menengah)": "#1f77b4", "🎯 Fokus Bimbingan (Bawah)": "#d62728"}
                )
                fig_ai.update_layout(title="Peta Sebaran Potensi Siswa Berbasis AI")
                st.plotly_chart(fig_ai, use_container_width=True)
            else:
                fig_ai = px.histogram(df, x="Rata-rata", color="Kelompok AI", hover_name="Nama", barmode="stack")
                st.plotly_chart(fig_ai, use_container_width=True)
                
            c_atas, c_tengah, c_bawah = st.columns(3)
            with c_atas:
                st.success("🌟 **Kelompok Akselerasi**")
                st.dataframe(df[df['Kelompok AI'] == "🌟 Akselerasi (Atas)"][['Nama', 'Rata-rata']].reset_index(drop=True), hide_index=True)
            with c_tengah:
                st.info("📈 **Kelompok Berkembang**")
                st.dataframe(df[df['Kelompok AI'] == "📈 Berkembang (Menengah)"][['Nama', 'Rata-rata']].reset_index(drop=True), hide_index=True)
            with c_bawah:
                st.error("🎯 **Fokus Bimbingan Intensif**")
                st.dataframe(df[df['Kelompok AI'] == "🎯 Fokus Bimbingan (Bawah)"][['Nama', 'Rata-rata']].reset_index(drop=True), hide_index=True)

        except Exception as e: st.warning(f"⚠️ Algoritma AI membutuhkan lebih banyak variasi data. Error: {e}")
    else: st.error("⚠️ Library scikit-learn tidak terinstal atau data nilai belum memadai.")

# --- TAB 5: SINKRONISASI KEDISIPLINAN (TEMPLATE MANUSIAWI) ---
with t5:
    st.header("🔄 Sinkronisasi Data Kedisiplinan & Ekstrakurikuler")
    st.info("💡 Langkah 2 (Opsional): Unggah sheet Data Personal/Absensi dari e-Rapor untuk mengaktifkan fitur Analisis Kedisiplinan.")
    
    with st.expander("📥 Unduh Template Dokumen BK & Absensi (Format Baru Berurutan - Manusiawi)", expanded=True):
        st.markdown("""
        **Tersedia 2 template instan otomatis berurutan (Bebas Kolom Kosong Tersembunyi):**
        - **Template Absensi & Ekskul (Terbaru)** → Kolom tersusun rapat berurutan tanpa jeda spasial (A sampai I) untuk meminimalisir salah ketik.
        - **Template Catatan BK** → Dokumen resmi rekap dossier bimbingan konseling siswa.
        """)
        
        nama_list = df['Nama'].tolist() if 'df_lengkap' in st.session_state else []
        kelas_for_template = st.session_state.get('nama_kelas', 'Kelas')
        
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.markdown("**📊 Template Absensi & Ekskul**")
            st.caption("Tata susunan kolom tertata rapat (A-I). Mencegah distorsi salin-tempel data!")
            if nama_list:
                template_absen_bytes = generate_template_absensi(nama_list, kelas_for_template)
                st.download_button(label="⬇️ Download Template Absensi", data=template_absen_bytes, file_name=f"Template_Absensi_Ekskul_{kelas_for_template}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else: st.button("⬇️ Download Template Absensi", disabled=True, use_container_width=True)
        
        with col_dl2:
            st.markdown("**📋 Template Catatan BK**")
            st.caption("Dokumen khusus rekapitualisasi laporan kasus perilaku konseling siswa.")
            if nama_list:
                template_bk_bytes = generate_template_bk(nama_list, kelas_for_template)
                st.download_button(label="⬇️ Download Template Catatan BK", data=template_bk_bytes, file_name=f"Template_Catatan_BK_{kelas_for_template}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else: st.button("⬇️ Download Template Catatan BK", disabled=True, use_container_width=True)
    
    st.markdown("---")
    st.markdown("**📤 Upload Data Absensi / BK yang Sudah Diisi:**")
    file_absen = st.file_uploader("Upload file Excel Rapor Tambahan (Absensi/BK):", type=['xlsx', 'csv'], key="absen_uploader")
    
    if file_absen:
        if hasattr(file_absen, 'seek'): file_absen.seek(0)
        try:
            excel_check = pd.read_excel(file_absen, sheet_name=0)
            file_absen.seek(0)
            cols = [str(c).strip().upper() for c in excel_check.columns]

            if all(c in cols for c in ['NAMA','H','S','I','A']):
                df_hsia = pd.read_excel(file_absen)
                df_hsia.columns = [str(c).strip().upper() for c in df_hsia.columns]
                df_hsia = df_hsia.rename(columns={'NAMA':'Nama','H':'Hadir','S':'Sakit','I':'Izin','A':'Alpa'})
                df_hsia['Nama'] = df_hsia['Nama'].astype(str).str.strip().str.upper()

                df_merge = df_utama.copy()
                df_merge['Nama_Upper'] = df_merge['Nama'].astype(str).str.strip().str.upper()
                df_merge = pd.merge(df_merge, df_hsia[['Nama','Hadir','Sakit','Izin','Alpa']], left_on='Nama_Upper', right_on='Nama', how='left').drop(columns=['Nama_y', 'Nama_Upper']).rename(columns={'Nama_x': 'Nama'})
                for col in ['Hadir','Sakit','Izin','Alpa']: df_merge[col] = pd.to_numeric(df_merge[col], errors='coerce').fillna(0)
                df_merge['Total_Hari'] = df_merge[['Hadir','Sakit','Izin','Alpa']].sum(axis=1)
                df_merge['Persentase_Hadir'] = ((df_merge['Hadir'] / df_merge['Total_Hari'].replace(0,1))*100).round(2)
                df_merge['Skor_Disiplin'] = df_merge['Persentase_Hadir'] - (df_merge['Alpa']*2)
                
                st.session_state.df_lengkap = df_merge
                st.session_state.absensi_synced = True
                st.success("✅ Data Rekap Kehadiran HSIA Berhasil Sinkron!")

            else:
                if file_absen.name.endswith('.csv'): df_absen = pd.read_csv(file_absen, header=None, skiprows=7)
                else: df_absen = pd.read_excel(file_absen, header=None, skiprows=7)

                # 🛠️ AMAN: Deteksi kata kunci global di 10 baris pertama untuk identifikasi template BK
                df_sample_bk = pd.read_excel(file_absen, header=None, nrows=10)
                all_text = " ".join(df_sample_bk.astype(str).values.flatten()).upper()
                
                if 'CATATAN BK' in all_text:
                    df_bk_clean = df_absen.copy()
                    df_bk_clean.columns = [str(i) for i in range(df_bk_clean.shape[1])]
                    df_bk_mapped = pd.DataFrame()
                    df_bk_mapped['NIS'] = df_bk_clean['2'].astype(str).str.replace(r'\.0$','',regex=True).str.strip()
                    df_bk_mapped['Catatan BK'] = df_bk_clean['10'].astype(str).str.strip()
                    df_bk_mapped['Jml Pelanggaran'] = pd.to_numeric(df_bk_clean['3'], errors='coerce').fillna(0)
                    
                    df_merged = pd.merge(df_utama, df_bk_mapped, on='NIS', how='left')
                    st.session_state.df_lengkap = df_merged
                    st.session_state.absensi_synced = True
                    st.success("✅ Sinkronisasi Berhasil: File Dokumen BK Sukses Dimuat.")
                else:
                    kolom_map = {2: 'NIS', 3: 'Sakit', 4: 'Izin', 5: 'Alpa', 6: 'Ekskul_Pramuka', 7: 'Ekskul_Kesenian', 8: 'Ekskul_Olahraga'}
                    df_absen = df_absen[[k for k in kolom_map.keys() if k in df_absen.columns]].rename(columns=kolom_map)
                    df_absen = df_absen[df_absen['NIS'].astype(str).str.upper() != 'NIS']

                    if 'NIS' in df_absen.columns:
                        df_absen['NIS'] = df_absen['NIS'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                        for c in ['Sakit', 'Izin', 'Alpa']: df_absen[c] = pd.to_numeric(df_absen[c], errors='coerce').fillna(0)
                        df_absen['Skor_Disiplin'] = 100 - (df_absen['Sakit']*1 + df_absen['Izin']*1 + df_absen['Alpa']*4)
                        df_merged = pd.merge(df_utama, df_absen, on='NIS', how='left')
                        st.session_state.df_lengkap = df_merged
                        st.session_state.absensi_synced = True
                        st.success("✅ Sinkronisasi Berhasil: File Template Absensi Berurutan Sukses Dimuat.")
                    try: st.rerun()
                    except AttributeError: st.experimental_rerun()
        except Exception as e: st.error(f"Terjadi kesalahan saat melakukan sinkronisasi data: {e}")

df = st.session_state.df_lengkap.copy()
if 'Total Nilai' in df.columns:
    df = df.sort_values(by='Total Nilai', ascending=False).reset_index(drop=True)
    df.index = df.index + 1

# --- TAB 6: DATABASE KELAS ---
with t6:
    st.header(f"📋 Database Nilai Leger {st.session_state.nama_kelas}")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_export = df.drop(columns=['Cluster_ID'], errors='ignore')
        df_export.to_excel(writer, index=False, sheet_name=st.session_state.nama_kelas)
    st.download_button(label="📥 Download Database Kelas (Excel)", data=buffer.getvalue(), file_name=f"Database_{st.session_state.nama_kelas}.xlsx")
    st.dataframe(df.drop(columns=['Cluster_ID'], errors='ignore'), use_container_width=True)